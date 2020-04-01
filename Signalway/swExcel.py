#!/usr/bin/python
# -*- coding: utf-8 -*-
# @File  : swExcel.py
# @Author: luowq
# @Date  : 2019/6/12 13:59
# @Desc  :
from openpyxl import  Workbook,load_workbook

class SwExcel(object):
    '''
    __init__: mode：模式 1: 读取 2：写入
    insertRows：按列插入
    insertColumn：按行插入
    save：保存文本生成文件
    '''
    def __init__(self,mode=2,file=""):
        if mode == 2:
            self.wb = Workbook()
            self.ws = self.wb.active
        else:
            self.wb = load_workbook(file)


    def getRowsToDict(self,sheet='Sheet1'):
        self.ws = self.wb[sheet]
        return self.ws

    def insertRows(self,datas):
        rows = list(zip(*datas))
        for row in rows:
            self.ws.append(row)


    def insertColumns(self,rows):
        for row in rows:
            self.ws.append(row)


    def save(self,name):
        self.wb.save(name)
        self.wb.close()


if __name__ == '__main__':
    excel = SwExcel(1,'../data/testCases.xlsx')
    datas = excel.getRowsToDict()
    varsName = []
    testDatas = []
    for index,d in enumerate(datas):
        if index == 0:
            varsName = [d[i].value for i in range(0,len(d))]
        else:
            print(dict(zip(varsName,[d[i].value for i in range(0,len(d))])))